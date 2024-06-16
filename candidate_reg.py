import tkinter as tk
from tkinter import ttk, messagebox, filedialog, simpledialog
from PIL import Image, ImageTk
from ttkbootstrap import Style
from openpyxl import Workbook, load_workbook
import os
import subprocess
import shutil

class CandidateManagementSystem:
    def __init__(self, root):
        self.root = root
        self.root.title("Candidate Management System")

        #self.style = Style("minty")
        self.root.attributes("-fullscreen", True)
        self.candidates = []
        self.admin_password = "admin123"
        self.ec_mode = False

        tk.Label(root, text="Candidate Management System", font=('Helvetica', 18, 'bold')).grid(row=0, column=1, columnspan=6, pady=10, sticky='nsew')

        self.create_add_delete_candidate_section()
        self.create_table_section()
        self.create_ec_view_modify_section()
        self.load_candidates_from_excel() 
        for i in range(8):
            self.root.grid_columnconfigure(i, weight=1)

        for i in range(13):
            self.root.grid_rowconfigure(i, weight=1)

    
    def save_candidates_to_excel(self): 
        wb = Workbook()
        ws = wb.active
        ws.append(["Party", "Party_Symbol", "Votes","Position","Name"])
        for candidate in self.candidates:
            ws.append(candidate)
        wb.save("sample.xlsx")

    def load_candidates_from_excel(self):
        if not os.path.exists("sample.xlsx"):
            self.save_candidates_to_excel()  # Create the file if it doesn't exist

        try:
            wb = load_workbook("sample.xlsx")
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                self.candidates.append(row)
            self.update_treeview()

        except Exception as e:
            messagebox.showerror("Error", f"An error occurred: {str(e)}")

    def create_add_delete_candidate_section(self):
        tk.Label(self.root, text="Candidate Name:", font=('Helvetica', 12)).grid(row=3, column=1, padx=5, pady=2, sticky=tk.E)
        self.name_entry = tk.Entry(self.root, font=('Helvetica', 12))
        self.name_entry.grid(row=3, column=2, padx=5, pady=2, columnspan=4, sticky=tk.W + 'nsew')

        tk.Label(self.root, text="Party Affiliation:", font=('Helvetica', 12)).grid(row=4, column=1, padx=5, pady=2, sticky=tk.E)
        self.party_entry = tk.Entry(self.root, font=('Helvetica', 12))
        self.party_entry.grid(row=4, column=2, padx=5, pady=2, columnspan=4, sticky=tk.W + 'nsew')

        tk.Label(self.root, text="Party Symbol", font=('Helvetica', 12)).grid(row=5, column=1, padx=5, pady=2, sticky=tk.E)
        self.bio_entry = tk.Entry(self.root, font=('Helvetica', 12))
        self.bio_entry.grid(row=5, column=2, padx=5, pady=2, columnspan=4, sticky=tk.W + 'nsew')

        tk.Label(self.root, text="Position:", font=('Helvetica', 12)).grid(row=6, column=1, padx=5, pady=2, sticky=tk.E)
        self.position_var = tk.StringVar()
        self.position_entry = ttk.Combobox(self.root, textvariable=self.position_var,
                                           values=["MLA", "MP"], font=('Helvetica', 12))
        self.position_entry.grid(row=6, column=2, padx=5, pady=2, columnspan=4, sticky=tk.W + 'nsew')

        tk.Label(self.root, text="Party Symbol:", font=('Helvetica', 12)).grid(row=8, column=1, padx=5, pady=2, sticky=tk.E)
        self.party_symbol_path = tk.StringVar()
        tk.Entry(self.root, textvariable=self.party_symbol_path, font=('Helvetica', 12)).grid(row=8, column=2, padx=5, pady=2, columnspan=3, sticky=tk.W + 'nsew')
        tk.Button(self.root, text="Browse", command=self.browse_party_symbol, font=('Helvetica', 12)).grid(row=8, column=5, padx=5, pady=2, sticky=tk.W + 'nsew')

        # Buttons
        tk.Button(self.root, text="Add Candidate", command=self.add_candidate, font=('Helvetica', 12)).grid(row=11, column=1,  padx=5, pady=2, sticky='nsew')
        tk.Button(self.root, text="Delete Candidate", command=self.delete_candidate, font=('Helvetica', 12)).grid(row=11, column=6, padx=5, pady=2, sticky='nsew')

    def create_table_section(self):
        table_frame = tk.Frame(self.root)
        table_frame.grid(row=12, column=1, columnspan=6, pady=5, sticky="nsew")

        self.candidate_tree = ttk.Treeview(table_frame, columns=("ID", "Name", "Position", "Party"), show="headings", style="mystyle.Treeview")
        self.candidate_tree.pack(side="left", fill="both", expand=True)

        scrollbar = tk.Scrollbar(table_frame, orient="vertical", command=self.candidate_tree.yview)
        scrollbar.pack(side="right", fill="y")

        self.candidate_tree.config(yscrollcommand=scrollbar.set)


        self.candidate_tree.heading("ID", text="ID", anchor=tk.CENTER)
        self.candidate_tree.heading("Name", text="Name", anchor=tk.CENTER)
        self.candidate_tree.heading("Position", text="Position", anchor=tk.CENTER)
        self.candidate_tree.heading("Party", text="Party", anchor=tk.CENTER)

        self.candidate_tree.column("ID", anchor=tk.CENTER, width=30)
        self.candidate_tree.column("Name", anchor=tk.CENTER, width=150)
        self.candidate_tree.column("Position", anchor=tk.CENTER, width=100)
        self.candidate_tree.column("Party", anchor=tk.CENTER, width=150)

        for i in range(1, 7):
            self.root.grid_columnconfigure(i, weight=1)
        self.root.grid_rowconfigure(10, weight=1)

    def exit_application(self):
        self.root.destroy()
        subprocess.run(["python", "main.py"])
    
    def create_ec_view_modify_section(self):
        self.mode_var = tk.IntVar()
        tk.Checkbutton(self.root, text="Election Commissioner Mode", variable=self.mode_var, command=self.toggle_ec_mode, font=('Helvetica', 12)).grid(row=13, column=1, columnspan=6, pady=5, sticky=tk.W)

        # Buttons
        tk.Button(self.root, text="View Details", command=self.view_details, font=('Helvetica', 12)).grid(row=15, column=1,  padx=4, pady=2, sticky='nsew')
        tk.Button(self.root, text="Modify Details", command=self.modify_details, font=('Helvetica', 12)).grid(row=15, column=6, padx=5, pady=2, sticky='nsew')

        tk.Button(self.root, text="Exit", command=self.exit_application, font=('Helvetica', 12),).grid(row=14, column=0, columnspan=8, pady=10)
        

        for i in range(1, 7):
            self.root.grid_columnconfigure(i, weight=1)
        self.root.grid_rowconfigure(11, weight=1)
        self.root.grid_rowconfigure(12, weight=1)

    def browse_party_symbol(self):
        file_path = filedialog.askopenfilename(filetypes=[("Image files", "*.png;*.jpg;*.jpeg")])
        self.party_symbol_path.set(file_path)

    def toggle_ec_mode(self):
        admin_password = simpledialog.askstring("Switch Mode", "Enter admin password:", show='*')

        if admin_password == self.admin_password:
            self.ec_mode = not self.ec_mode
            if self.ec_mode:
                messagebox.showinfo("Election Commissioner Mode", "Election Commissioner mode activated.\n")
        else:
            messagebox.showwarning("Invalid Password", "Incorrect admin password. Switching to Candidate mode.")
            self.ec_mode = False
            self.mode_var.set(0)

    def add_candidate(self):
        if self.ec_mode:
            name = self.name_entry.get()
            party = self.party_entry.get()
            symbol = self.bio_entry.get()
            position = self.position_var.get()
            party_symbol = self.party_symbol_path.get()
            if name and party and symbol and position and party_symbol:
                if self.authenticate_ec():
                    if position == "MLA":
                        party_symbol_folder = "Party_symbols"
                    elif position == "MP":
                        party_symbol_folder = "Party_symbols1"
                    else:
                        messagebox.showwarning("Invalid Position", "Please select a valid position (MLA or MP).")
                        return
                    os.makedirs(party_symbol_folder, exist_ok=True)
                    party_symbol_filename = os.path.basename(party_symbol)
                    new_party_symbol_path = os.path.join(party_symbol_folder, f"{party}.png")
                    shutil.copy(party_symbol, new_party_symbol_path)
                    candidate_info = (party, symbol, 0, position, name )
                    self.candidates.append(candidate_info)
                    self.update_treeview()
                    self.clear_entries()
                else:
                    messagebox.showwarning("Invalid Credential", "Incorrect password.")
            else:
                messagebox.showwarning("Incomplete Information", "Please fill in all fields.")
            self.save_candidates_to_excel()
        else:
            messagebox.showwarning("Access Denied!","Please select ec mode.")

    def delete_candidate(self):
        if self.ec_mode:
            if self.authenticate_ec():            
                selected_index = self.candidate_tree.selection()
                if selected_index:
                    deleted_candidate = self.candidates.pop(int(selected_index[0]) - 1)
                    party, _, _, position, _ = deleted_candidate
                    if position == "MLA":
                        party_symbol_folder = "Party_symbols"
                    else:
                        party_symbol_folder = "Party_symbols1"
                    os.makedirs(party_symbol_folder, exist_ok=True)
                    new_party_symbol_path = os.path.join(party_symbol_folder, f"{party}.png")
                    if os.path.exists(new_party_symbol_path):
                        os.remove(new_party_symbol_path)
                    messagebox.showinfo("Deleted", f"Candidate {deleted_candidate[0]} has been deleted.")
                    self.update_treeview()
                else:
                    messagebox.showwarning("No Candidate Selected", "Please select a candidate to delete.")
            self.save_candidates_to_excel()
        else:
            messagebox.showwarning("Access Denied!","Please select ec mode.")        

    def view_details(self):
        selected_index = self.candidate_tree.selection()
        if selected_index:
            index_as_int = int(selected_index[0])
            candidate = self.candidates[index_as_int - 1]
            self.show_candidate_details(candidate)
        else:
            messagebox.showwarning("No Candidate Selected", "Please select a candidate to view details.")


    def show_candidate_details(self, candidate):
        details_window = tk.Toplevel(self.root)
        details_window.title("Candidate Details")
        details_window.attributes('-toolwindow', True)
        details_window.resizable(width=False, height=False)
        
        tk.Label(details_window, text="Party Name:").grid(row=0, column=0, padx=10, pady=5, sticky=tk.E)
        tk.Label(details_window, text="Party Affiliation:").grid(row=1, column=0, padx=10, pady=5, sticky=tk.E)
        tk.Label(details_window, text="Position:").grid(row=2, column=0, padx=10, pady=5, sticky=tk.E)
        tk.Label(details_window, text="Name:").grid(row=3, column=0, padx=10, pady=5, sticky=tk.E)
        tk.Label(details_window, text="Symbol:").grid(row=4, column=0, padx=10, pady=5, sticky=tk.E)

        tk.Label(details_window, text=candidate[0]).grid(row=0, column=1, padx=10, pady=5, sticky=tk.W)
        tk.Label(details_window, text=candidate[1]).grid(row=1, column=1, padx=10, pady=5, sticky=tk.W)
        tk.Label(details_window, text=candidate[3]).grid(row=2, column=1, padx=10, pady=5, sticky=tk.W)
        tk.Label(details_window, text=candidate[4]).grid(row=3, column=1, padx=10, pady=5, sticky=tk.W)

        if candidate[3] == 'MLA':
            img = './Party_Symbols/'+ candidate[0] + '.png'
        elif candidate[3] == 'MP':
            img = './Party_Symbols1/'+ candidate[0] + '.png'
        else:
            img = 'broken.png'

        party_symbol = Image.open(img)
        party_symbol = party_symbol.resize((100, 100))
        party_symbol = ImageTk.PhotoImage(party_symbol)
        tk.Label(details_window, image=party_symbol).grid(row=5, column=1, padx=10, pady=5, sticky=tk.W)
        tk.Label(details_window, image=party_symbol).image = party_symbol
        details_window.wait_window()

    def modify_details(self):
        if self.ec_mode:
            if self.authenticate_ec():
                selected_index = self.candidate_tree.selection()
            
                if selected_index:
                    self.modify_candidate_details(int(selected_index[0]) - 1)
                else:
                    messagebox.showwarning("No Candidate Selected", "Please select a candidate to modify.")
            self.save_candidates_to_excel()
        else:
            messagebox.showwarning("Access Denied!","Please select ec mode.")

    def browse_party_symbol_modify(self):
        file_path = filedialog.askopenfilename(filetypes=[("Image files", "*.png;*.jpg;*.jpeg")])
        return file_path
    
    def modify_candidate_details(self, index, path=''):
        details_window = tk.Toplevel(self.root)
        details_window.title("Modify Details")
        details_window.attributes('-toolwindow', True)
        details_window.resizable(width=False, height=False)

        tk.Label(details_window, text="Party Name:").grid(row=0, column=0, padx=10, pady=5, sticky=tk.E)
        name_entry = tk.Entry(details_window)
        name_entry.grid(row=0, column=1, padx=10, pady=5, columnspan=2)
        name_entry.insert(0, self.candidates[index][0])
 
        tk.Label(details_window, text="Party Affiliation:").grid(row=1, column=0, padx=10, pady=5, sticky=tk.E)
        party_entry = tk.Entry(details_window)
        party_entry.grid(row=1, column=1, padx=10, pady=5, columnspan=3)
        party_entry.insert(0, self.candidates[index][1])

        tk.Label(details_window, text="Position:").grid(row=3, column=0, padx=10, pady=5, sticky=tk.E)
        position_var = tk.StringVar()
        position_entry = ttk.Combobox(details_window, textvariable=position_var, values=["MP", "MLA"])
        position_entry.grid(row=3, column=1, padx=10, pady=5, columnspan=3, sticky=tk.W)
        position_entry.set(self.candidates[index][3])

        tk.Label(details_window, text="Party Symbol:").grid(row=5, column=0, padx=10, pady=5, sticky=tk.E)
        party_symbol_path_var = tk.StringVar()
        tk.Entry(details_window, textvariable=party_symbol_path_var).grid(row=5, column=1, padx=10, pady=5, columnspan=2)
        
        def browse():
            path = self.browse_party_symbol_modify()
            self.modify_candidate_details(index, path)
            #party_symbol_path_var.set(path)

        tk.Button(details_window, text="Browse", command=lambda: browse()).grid(row=5, column=3, padx=10, pady=5)
        def save_action():
            new_name = name_entry.get()
            new_party = party_entry.get()
            new_position = position_var.get()
            new_symbol_path = party_symbol_path_var.get()
            new_symbol_path = self.symbol_modify(new_symbol_path, new_position, new_name, old_img)
            self.save_changes(index, new_name, new_party, new_position)

        tk.Button(details_window, text="Save Changes", command=lambda: save_action(), font=('Helvetica', 12)).grid(row=6, column=0, columnspan=4, pady=10)
        if self.candidates[index][3] == "MLA":
            party_symbol_folder = "Party_symbols"
        elif self.candidates[index][3] == "MP":
            party_symbol_folder = "Party_symbols1"
        else:
            messagebox.showwarning("Invalid Position", "Please select a valid position (MLA or MP).")
            return
        
        current_directory = os.getcwd()
        old_img = current_directory+"\\"+party_symbol_folder+f"\\{name_entry.get()}.png"
        if path:
            party_symbol_path_var.set(path)
        else:
            party_symbol_path_var.set(old_img)

    
    def symbol_modify(self, party_symbol_path, position, name, old_img):
        if position == "MLA":
            party_symbol_folder = "Party_symbols"
        elif position == "MP":
            party_symbol_folder = "Party_symbols1"
        else:
            raise ValueError("Invalid position. Must be 'MP' or 'MLA'.")

        os.makedirs(party_symbol_folder, exist_ok=True)
        current_directory = os.getcwd()
        new_party_symbol_path = current_directory+"\\"+party_symbol_folder+f"\\{name}.png"
        src_norm = os.path.normcase(party_symbol_path)
        dst_norm = os.path.normcase(new_party_symbol_path)
        if src_norm == dst_norm:
            os.rename(party_symbol_path, new_party_symbol_path)
            return new_party_symbol_path
        else:
            shutil.copy(party_symbol_path, new_party_symbol_path)     
            src_norm = os.path.normcase(old_img) 
            if src_norm != dst_norm:
                if os.path.exists(old_img):
                    os.remove(old_img)
            return new_party_symbol_path

    def photo_modify(self, photo_path_var):
        file_path = filedialog.askopenfilename(filetypes=[("Image files", "*.png;*.jpg;*.jpeg")])
        photo_path_var.set(file_path)

    
    
    def save_changes(self, index, name, party, position):
        if name and party and position:
            self.candidates[index] = (name, party, self.candidates[index][2], position, self.candidates[index][4])
            messagebox.showinfo("Changes Saved", "Candidate details have been successfully modified.")
            self.update_treeview()
            self.save_candidates_to_excel()
        else:
            messagebox.showwarning("Incomplete Information", "Please fill in all fields.")

    def update_treeview(self):
        for item in self.candidate_tree.get_children():
            self.candidate_tree.delete(item)

        for i, candidate_info in enumerate(self.candidates, start=1):
            self.candidate_tree.insert("", tk.END, iid=i, values=(i, candidate_info[0], candidate_info[3], candidate_info[1]))

    def clear_entries(self):
        self.name_entry.delete(0, tk.END)
        self.party_entry.delete(0, tk.END)
        self.bio_entry.delete(0, tk.END)
        self.position_entry.set("")
        self.party_symbol_path.set("")

    def authenticate_ec(self):
        password = simpledialog.askstring("Election Commissioner Mode", "Enter admin password:", show='*')
        if password == self.admin_password:
            return True
        else:
            messagebox.showwarning("Invalid Credentials", "Incorrect password.")
            return False


if __name__ == "__main__":
    root = tk.Tk()
    app = CandidateManagementSystem(root)
    root.mainloop()

