import os
import tkinter as tk
from tkinter import messagebox
from PIL import Image, ImageTk
import openpyxl
import subprocess

class ImageApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Voting Interface Module")
        self.root.attributes("-fullscreen", True)

        self.folder_path = os.path.join(os.path.dirname(__file__), "Party_Symbols1")

        self.image_files = [f for f in os.listdir(self.folder_path) if f.endswith('.png')]

        self.load_data()

        self.create_widgets()
        self.bind_mouse_scroll()  # Bind mouse scroll event after canvas creation

    def load_data(self):
        self.party_names = []
        self.vote_counts = []

        try:
            wb = openpyxl.load_workbook("sample.xlsx")
            ws = wb.active

            for row in ws.iter_rows(min_row=2, max_col=3, values_only=True):
                self.party_names.append(row[0])
                self.vote_counts.append(row[2])

        except FileNotFoundError:
            messagebox.showerror("Error", "Votes data file not found.")

    def update_data(self, party_name):
        idx = self.party_names.index(party_name)
        self.vote_counts[idx] = int(self.vote_counts[idx])+ 1

        wb = openpyxl.load_workbook("sample.xlsx")
        ws = wb.active

        for i, (party_name, vote_count) in enumerate(zip(self.party_names, self.vote_counts), start=2):
            if(ws.cell(row=i,column=4).value=='MP'):
                ws.cell(row=i, column=3).value = vote_count

        wb.save("sample.xlsx")

    def create_widgets(self):
        header_label = tk.Label(self.root, text="Choose Political Party : MP", font=("Helvetica", 18), pady=20, fg="blue")
        header_label.pack()

        canvas = tk.Canvas(self.root, height=self.root.winfo_screenheight())
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        self.canvas = canvas  # Store canvas as an attribute of the class

        scrollbar = tk.Scrollbar(self.root, command=canvas.yview)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        canvas.configure(yscrollcommand=scrollbar.set)

        frame = tk.Frame(canvas)
        canvas.create_window((0, 0), window=frame, anchor=tk.NW)

        frame_width = 300
        frame_height = 300

        widgets_in_row = 4
        row_count = 0
        for i, file_name in enumerate(self.image_files):
            image_path = os.path.join(self.folder_path, file_name)
            img = Image.open(image_path)
            img = img.resize((frame_width - 20, frame_height - 50))
            img = ImageTk.PhotoImage(img)

            padx_value = 40
            pady_value = 30
            column_value = i % widgets_in_row

            button = tk.Button(frame, image=img, text=file_name.split('.')[0], compound=tk.TOP, command=lambda f=file_name: self.button_click(f))
            button.image = img
            button["font"] = ("Helvetica", 12)
            button.grid(row=row_count, column=column_value, padx=padx_value, pady=pady_value)

            if (i + 1) % widgets_in_row == 0:
                row_count += 1

        frame.update_idletasks()
        canvas.config(scrollregion=canvas.bbox("all"))

    def bind_mouse_scroll(self):
        self.root.bind("<MouseWheel>", self.on_mouse_scroll)

    def on_mouse_scroll(self, event):
        canvas_height = self.root.winfo_screenheight()
        if event.y < canvas_height:
            self.canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

    def button_click(self, file_name):
        print(f"Button clicked for {file_name}")

        party_name = file_name.split('.')[0]
        self.update_data(party_name)

        for widget in self.root.winfo_children():
            widget.destroy()

        label = tk.Label(self.root, text="Thanks for voting!", font=("Helvetica", 24), pady=50)
        label.pack(fill=tk.BOTH, expand=True)
        
        self.root.after(5000, self.redirect_to_login)
        
    def redirect_to_login(self):
        self.root.destroy()
        # Open Voting_Inteface_Module.py
        subprocess.run(["python", "log.py"])

if __name__ == "__main__":
    root = tk.Tk()
    app = ImageApp(root)
    root.mainloop()
