from tkinter import *
import tkinter as tk
from tkinter import messagebox,simpledialog
import ttkbootstrap as tb
import datetime
from PIL import Image, ImageTk
import openpyxl
import subprocess

def check_credentials(election_id, date_of_birth):
    try:
        dob_datetime = datetime.datetime.strptime(date_of_birth, '%Y/%m/%d')
    except ValueError:
        return False, None

    wb = openpyxl.load_workbook("userdata.xlsx")
    ws = wb.active
    found = False
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row[1] is None:
            continue
        excel_dob = row[1].strftime('%Y/%m/%d')
        if str(row[0]).strip() == str(election_id).strip() and excel_dob == date_of_birth:
            flag_cell = ws.cell(row=idx, column=3)
            flag = flag_cell.value
            if flag == 0:
                flag_cell.value = 1
                wb.save("userdata.xlsx")
                return True, wb  # Return the workbook to save later
            else:
                return False, None
        found = True
    if not found:
        wb.save("userdata.xlsx")
    return False, None

def checker():
    election_id = Election_ID.get()
    date_of_birth = DOB.get()

    # Check if fields are empty
    if election_id.strip() == "" or date_of_birth.strip() == "":
        login_label.config(text="All fields must be filled before login", fg="red")
        return
    success, wb = check_credentials(election_id, date_of_birth)
    if success:
        login_label.config(text="Login Successful", fg="green")
        # Save the workbook here
        if wb:
            wb.save("userdata.xlsx")
        root.destroy()
        # Open Voting_Inteface_Module.py
        subprocess.run(["python", "Voting_Interface_Module.py"])
    else:
        login_label.config(text="Invalid credentials", fg="red")

def on_entry_click(event, entry_widget, default_text):
    if entry_widget.get() == default_text:
        entry_widget.delete(0, tk.END)
        entry_widget.config(fg='black')

def on_focus_out(event, entry_widget, default_text):
    if entry_widget.get() == "":
        entry_widget.insert(0, default_text)
        entry_widget.config(fg='grey')

def exit_application():
        pswd =  "admin123"
        admin_password = simpledialog.askstring("Enter Password","Enter admin password:", show='*')
        if admin_password == pswd:
            root.destroy()
            subprocess.run(["python", "main.py"])
        else:
            messagebox.showwarning("Invalid Password", "Incorrect admin password. Can't exit.")
        

root = tb.Window(themename="litera", iconphoto=None)
root.configure(bg="#f5f5f5")
root.attributes('-fullscreen', True)
root.title("Log In ")

registration_label = tb.Label(root, text="Login", font=("Roboto", 40, "bold"))
registration_label.pack(pady=(50,0))
registration_label.configure(style="Primary.TLabel")

my_frame = LabelFrame(root ,font=("Roboto", 16, "bold"), background="#00FFFF", padx=100, pady=100)
my_frame.place(relx=0.5, rely=0.5, anchor='center')

Election_ID = tk.StringVar()
DOB = tk.StringVar()

entry_election_id = tk.Entry(my_frame, font=("Roboto", 16), width=30, fg='grey', textvariable=Election_ID)
entry_election_id.insert(0, "Election ID")
entry_election_id.bind("<FocusIn>", lambda event: on_entry_click(event, entry_election_id, "Election ID"))
entry_election_id.bind("<FocusOut>", lambda event: on_focus_out(event, entry_election_id, "Election ID"))
entry_election_id.pack(pady=15, padx=20)

entry_dob = tk.Entry(my_frame, font=("Roboto", 16), width=30, fg='grey', textvariable=DOB)
entry_dob.insert(0, "Date of Birth (YYYY/MM/DD)")
entry_dob.bind("<FocusIn>", lambda event: on_entry_click(event, entry_dob, "Date of Birth (YYYY/MM/DD)"))
entry_dob.bind("<FocusOut>", lambda event: on_focus_out(event, entry_dob, "Date of Birth (YYYY/MM/DD)"))
entry_dob.pack(pady=15, padx=20)

my_button = tb.Button(my_frame, text="Log In →", bootstyle="success", command=checker)
my_button.pack(pady=(60,0), padx=20)

exit_button = tb.Button(my_frame, text="Exit", bootstyle="danger", command=exit_application)
exit_button.pack(pady=10, padx=20)


login_label = Label(my_frame, font=("Roboto", 16))
login_label.pack(pady=(20,0))


original_image = Image.open("emblem.png")
resized_image = original_image.resize((90, 90))
logo_image = ImageTk.PhotoImage(resized_image)

logo_label = tb.Label(root, image=logo_image, anchor='s')
logo_label.place(relx=0.5, rely=1.0, anchor='s', y=-30)

root.option_add("*Font", "Roboto")
root.option_add("*Entry.highlightcolor", "#54a0ff")

root.mainloop()